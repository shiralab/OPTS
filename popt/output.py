import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path

import openpyxl
from openpyxl.styles import Font


class ResultWriter:
    def __init__(self, path, sheet, title=""):
        self.font = Font(name="メイリオ", size=12)
        # self.alignment = Alignment(wrap_text=True)
        self.sheet = sheet
        if isinstance(path, str):
            self.path = Path(path)
        else:
            self.path = path

        if not self.path.parent.exists():
            self.path.parent.mkdir(parents=True)

        file = openpyxl.Workbook()
        file_sheet = file.active
        file_sheet.title = sheet
        file.save(self.path)

        if title != "":
            file = openpyxl.load_workbook(self.path)
            write_sheet = file[self.sheet]
            write_row = 1
            write_sheet.cell(row=write_row, column=1).value = title

            file.save(self.path)

    def record2excel(
        self, step, parents, offspring, ptr, applied_offspring, score, increse
    ):
        record = []
        record.extend(parents)
        record.append(offspring)
        record.append(ptr)
        record.append(applied_offspring)
        record.append(score)
        record.append(increse)

        sheet = f"step{step}_way"
        header = []
        for i in range(len(parents)):
            header.append(f"parents{i + 1}")
        header.append("offspring")
        header.append("ptr")
        header.append("applied offspring")
        header.append("score")
        header.append("increse")

        self.write2excel(sheet=sheet, header=header, contents=[record])

    def write2excel(self, sheet="default", title="", header=[], contents=[]):
        new_sheet = False
        file = openpyxl.load_workbook(self.path)
        if sheet == "default":
            write_sheet_name = self.sheet
        else:
            write_sheet_name = sheet

        file_sheets = file.sheetnames
        if write_sheet_name not in file_sheets:
            file.create_sheet(title=write_sheet_name)
            new_sheet = True

        write_sheet = file[write_sheet_name]
        write_row = write_sheet.max_row + 1

        if title != "":
            write_sheet.cell(row=write_row, column=1).value = title
            write_row += 1

        if header and new_sheet:
            for i, elem in enumerate(header):
                write_sheet.cell(row=write_row, column=i + 1).value = elem
            write_row += 1

        for i, elem in enumerate(contents):
            if isinstance(elem, dict):
                for j, key in enumerate(elem):
                    write_sheet.cell(row=write_row + i, column=j + 1).value = elem[key]
                    write_sheet.cell(row=write_row + i, column=j + 1).font = self.font
            elif isinstance(elem, list):
                for j, val in enumerate(elem):
                    write_sheet.cell(row=write_row + i, column=j + 1).value = val
                    write_sheet.cell(row=write_row + i, column=j + 1).font = self.font

        file.save(self.path)


def setup_log(log_path, log_name="default", stream=True):
    print("Setting up log for", log_name)
    logger = logging.getLogger(log_name)
    if not logger.handlers:
        logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter("[%(asctime)s] [%(levelname)-5s] - %(message)s")
        file_handler = RotatingFileHandler(
            filename=log_path, maxBytes=10000000, backupCount=30
        )
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        if stream:
            stream_handler = logging.StreamHandler()
            stream_handler.setFormatter(formatter)
            logger.addHandler(stream_handler)
    return logger
